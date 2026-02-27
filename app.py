import threading
import time
import shutil
from fastapi import Cookie
from fastapi import Form
from fastapi.responses import RedirectResponse, StreamingResponse
import hashlib
import json
import os, re, datetime, uuid, threading, decimal, io
from urllib.parse import quote
from typing import List, Optional, Dict
from fastapi import FastAPI, Header, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import pandas as pd

app = FastAPI(title="PDF Search (auth on)")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

CONFIG_LOCK = threading.Lock()
CONFIG: Dict = {
    "roots": ["/data/contracts/2024", "/data/contracts/2025", "/data/contracts/2021", "/data/contracts/2022", "/data/contracts/2023", "/data/contracts/2000", "/data/contracts/2001", "/data/contracts/2002", "/data/contracts/2003", "/data/contracts/2004", "/data/contracts/2005", "/data/contracts/2006", "/data/contracts/2007", "/data/contracts/2008", "/data/contracts/2009", "/data/contracts/2010", "/data/contracts/2011", "/data/contracts/2012", "/data/contracts/2013", "/data/contracts/2014", "/data/contracts/2015", "/data/contracts/2016", "/data/contracts/2017", "/data/contracts/2018", "/data/contracts/2019", "/data/contracts/2020", "/data/contracts/2026", "/data/contracts/2027", "/data/contracts/2028", "/data/contracts/2029", "/data/contracts/2030", "/data/contracts/2031", "/data/contracts/2032", "/data/contracts/2033", "/data/contracts/2034", "/data/contracts/2035", "/data/contracts/2036", "/data/contracts/2037", "/data/contracts/2038", "/data/contracts/2039", "/data/contracts/2040", "/data/contracts/2041", "/data/contracts/2042", "/data/contracts/2043", "/data/contracts/2044", "/data/contracts/2045", "/data/contracts/2046", "/data/contracts/2047", "/data/contracts/2048", "/data/contracts/2049", "/data/contracts/2050"],
    "excel_patterns": ["INDEX.XLSX"],      # 文件名大小写不敏感
    "allowed_exts": [".pdf"],

    # 公网映射（与 Nginx /files /dl 对应）
    "public_base": "/data/contracts",
    "preview_prefix": "/files/",
    "download_prefix": "/dl/",

    # PDF 子目录（年份下的 DOCS 或 docs）
    "pdf_subdirs": ["DOCS", "docs"],

    # 行为开关（默认满足你的当前要求）
    "case_insensitive": True,             # 文本大小写不敏感
    "amount_numeric_equivalence": True,   # 金额数值等价
    "amount_zero_means_empty": True,      # 金额为0视为未填（忽略）
    "text_logic_or": False,               # 文本条件采用 AND（多条件都要命中）
}

RETURN_FIELDS = ["序号","工程地点及内容","单位名称","签订日期","合同额","结算值","已付款","欠付款","合同编号","pdf_path","pdf_dl"]

PASSWORD_FALLBACK = "1982567"
PASSWORDS_FILE = os.path.join(os.path.dirname(__file__), "DATA", "passwords.txt")
TOKENS: Dict[str, float] = {}

APP_DIR = os.path.dirname(__file__)

AUTO_UPDATE_LOCK = threading.Lock()
AUTO_UPDATE_ENABLED = False
PENDING_DIR = "/www/wwwroot/dav.gfwzb.com/data"
ADD_PENDING_DIR = "/data/add_pending"

def _load_passwords() -> set[str]:
    try:
        with open(PASSWORDS_FILE, "r", encoding="utf-8") as f:
            raw = f.read()
    except FileNotFoundError:
        return {PASSWORD_FALLBACK}
    passwords = {p.strip() for p in raw.split(",") if p.strip()}
    return passwords or {PASSWORD_FALLBACK}

def _password_valid(password: str) -> bool:
    return str(password) in _load_passwords()

def _resolve_data_dir() -> str:
    lower = os.path.join(APP_DIR, "data")
    upper = os.path.join(APP_DIR, "DATA")
    if os.path.isdir(lower):
        return lower
    return upper

def _auto_update_log(message: str) -> None:
    data_dir = _resolve_data_dir()
    os.makedirs(data_dir, exist_ok=True)
    log_path = os.path.join(data_dir, "auto_update.log")
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {message}\n"
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(line)

def _load_auto_update_enabled() -> bool:
    data_dir = _resolve_data_dir()
    path = os.path.join(data_dir, "auto_update.json")
    try:
        with open(path, "r", encoding="utf-8") as f:
            payload = json.load(f)
            return bool(payload.get("enabled", False))
    except FileNotFoundError:
        return False
    except Exception:
        return False

def _save_auto_update_enabled(enabled: bool) -> None:
    data_dir = _resolve_data_dir()
    os.makedirs(data_dir, exist_ok=True)
    path = os.path.join(data_dir, "auto_update.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump({"enabled": bool(enabled)}, f, ensure_ascii=False, indent=2)

def _load_run_state() -> Dict[str, str]:
    data_dir = _resolve_data_dir()
    path = os.path.join(data_dir, "auto_update_runs.json")
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
            return {
                "copy_date": data.get("copy_date", ""),
                "update_date": data.get("update_date", ""),
            }
    except FileNotFoundError:
        return {"copy_date": "", "update_date": ""}
    except Exception:
        return {"copy_date": "", "update_date": ""}

def _save_run_state(state: Dict[str, str]) -> None:
    data_dir = _resolve_data_dir()
    os.makedirs(data_dir, exist_ok=True)
    path = os.path.join(data_dir, "auto_update_runs.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)

def _set_auto_update_enabled(enabled: bool) -> None:
    global AUTO_UPDATE_ENABLED
    with AUTO_UPDATE_LOCK:
        AUTO_UPDATE_ENABLED = bool(enabled)
        _save_auto_update_enabled(AUTO_UPDATE_ENABLED)

def _get_auto_update_enabled() -> bool:
    with AUTO_UPDATE_LOCK:
        return bool(AUTO_UPDATE_ENABLED)

# 轻量 rows 缓存（按 Excel 列表+mtime 签名）
_ROWS_CACHE = {"sig": None, "rows": []}
_ROWS_LOCK = threading.Lock()

def _utc_now_ts() -> float: return datetime.datetime.utcnow().timestamp()
def _issue_token(hours=24) -> str:
    t = str(uuid.uuid4()); TOKENS[t] = (_utc_now_ts() + hours*3600); return t
def require_auth(x_auth: str = Header(None), X_Auth: str = Cookie(None)):
    tok = x_auth or X_Auth
    if not tok or tok not in TOKENS:
        raise HTTPException(401, "Unauthorized")
    if TOKENS[tok] < _utc_now_ts():
        TOKENS.pop(tok, None)
        raise HTTPException(401, "Token expired")
    return tok

def _norm_date(s:str)->str:
    """
    统一日期为 YYYY[-MM[-DD]]
    兼容：中文年月日、全角数字、任意分隔符(./-_/空格)、带时间、连写(YYYYMM/ YYYYMMDD)、Excel 序列号
    """
    if s is None: return ""
    t = str(s).strip()
    if t == "": return ""

    # 去掉时间部分（空格后）
    t = t.split(' ')[0]

    # 全角->半角数字
    def to_halfwidth(u):
        return ''.join(chr(ord(ch)-0xFEE0) if '０'<=ch<='９' else ch for ch in u)
    t = to_halfwidth(t)

    # 统一分隔符
    t = t.replace('年','-').replace('月','-').replace('日','')
    t = re.sub(r'[._/\\\s]+','-', t)

    # 纯数字优先：YYYY / YYYYMM / YYYYMMDD（避免误判为 Excel 序列号）
    if re.fullmatch(r'\d+(?:\.\d+)?', t):
        try:
            num = float(t)
            if num.is_integer():
                num_int = int(num)
                num_str = str(num_int)
                if len(num_str) == 4 and 1900 <= num_int <= 2100:
                    return num_str
                if len(num_str) == 6:          # YYYYMM
                    return num_str[:4] + '-' + num_str[4:6]
                if len(num_str) == 8:          # YYYYMMDD
                    return num_str[:4] + '-' + num_str[4:6] + '-' + num_str[6:8]
            if num >= 1:
                base = datetime.datetime(1899, 12, 30)
                dt = base + datetime.timedelta(days=num)
                return dt.strftime('%Y-%m-%d')
        except Exception:
            pass

    m = re.match(r'^(\d{4})(?:-(\d{1,2}))?(?:-(\d{1,2}))?', t)
    if not m: return ""
    y = m.group(1)
    mo = m.group(2).zfill(2) if m.group(2) else ""
    d  = m.group(3).zfill(2) if m.group(3) else ""
    return "-".join([x for x in [y,mo,d] if x])

def _iter_excel_files(root: str):
    """在指定 root 下按 excel_patterns 精确匹配文件名（大小写不敏感），直接返回该层的 Excel 文件。"""
    want = [p.lower() for p in CONFIG.get("excel_patterns", [])]
    try:
        for name in os.listdir(root):
            p = os.path.join(root, name)
            if os.path.isfile(p) and name.lower() in want:
                yield p
    except FileNotFoundError:
        return

def _gather_excel_files():
    """返回 (files_in_order, signature)：
    files_in_order：按 CONFIG["roots"] 与 _iter_excel_files 顺序收集到的 Excel 路径
    signature：对 (Excel清单: 路径,mtime_ns,size) + (PDF目录清单: 路径,mtime_ns,项数) + excel_patterns 做摘要
    """
    files=[]    # [(excel_path, mtime_ns, size)]
    dirs=[]     # [(dir_path, mtime_ns, count)]
    pdf_subdirs = CONFIG.get("pdf_subdirs", [])
    for root in CONFIG.get("roots", []):
        if not os.path.isdir(root):
            continue
        # 目录签名：root 以及其 pdf_subdirs（存在才统计）
        for d in [root] + [os.path.join(root, sd) for sd in (pdf_subdirs or [])]:
            if os.path.isdir(d):
                try:
                    st = os.stat(d)
                    mns = int(getattr(st, "st_mtime_ns", int(st.st_mtime*1e9)))
                    try:
                        cnt = len(os.listdir(d))
                    except Exception:
                        cnt = -1
                    dirs.append((d, mns, cnt))
                except Exception:
                    dirs.append((d, 0, -1))
        # Excel 清单
        for x in _iter_excel_files(root):
            try:
                st = os.stat(x)
                mns = int(getattr(st, "st_mtime_ns", int(st.st_mtime*1e9)))
                files.append((x, mns, st.st_size))
            except Exception:
                files.append((x, 0, 0))
    # 稳定排序后摘要
    sig_files = sorted(files, key=lambda t: t[0].lower())
    sig_dirs  = sorted(dirs,  key=lambda t: t[0].lower())
    payload = json.dumps({
        "files": sig_files,
        "dirs":  sig_dirs,
        "excel_patterns": CONFIG.get("excel_patterns", [])
    }, ensure_ascii=False, sort_keys=True)
    sig = hashlib.sha1(payload.encode("utf-8")).hexdigest()
    return [f[0] for f in files], sig
def _normalize_amount_to_decimal(s: str):
    """金额数值等价（去货币符号/千分位/空格；仅保留0-9.-）；0 视为未填（可配置）"""
    if s is None: return None
    raw=str(s).strip()
    if raw=="": return None
    t=(raw.replace("￥","").replace("¥","").replace("人民币","").replace("元","")
         .replace("CNY","").replace("RMB","").replace(",","").replace(" ",""))
    t=re.sub(r"[^0-9.\-]","",t)
    if t=="":
        return None
    try:
        val=decimal.Decimal(t)
    except Exception:
        return None
    if CONFIG.get("amount_zero_means_empty", True) and val==0:
        return None
    return val

def _find_pdf(root, base, exts, subdirs):
    """在 root、root/subdir 下找 PDF：完整等于 base.ext；或“以 base 开头”的 pdf；大小写不敏感"""
    if not base: return ""
    base = str(base).strip()
    if not base: return ""
    cand_dirs = [root]
    for sd in (subdirs or []):
        cand_dirs.append(os.path.join(root, sd))
    for d in cand_dirs:
        try:
            # 1) 完全等于 base.ext（大小写忽略）
            for ext in exts:
                p=os.path.join(d, f"{base}{ext}")
                if os.path.isfile(p): return p
                low=f"{base}{ext}".lower()
                for name in os.listdir(d):
                    if name.lower()==low and name.lower().endswith(ext.lower()):
                        q=os.path.join(d, name)
                        if os.path.isfile(q): return q
            # 2) 以 base 开头的 *.pdf
            for name in os.listdir(d):
                ln=name.lower()
                if ln.endswith(".pdf") and ln.startswith(base.lower()):
                    q=os.path.join(d, name)
                    if os.path.isfile(q): return q
        except FileNotFoundError:
            continue
    return ""

def _load_all_rows():
    # —— 轻量缓存：按 Excel 清单 + (mtime,size) 签名 —— 
    files, sig = _gather_excel_files()
    with _ROWS_LOCK:
        if _ROWS_CACHE.get("sig") == sig:
            return list(_ROWS_CACHE.get("rows", []))

    rows=[]
    rows_by_key={}
    for x in files:
        scan_root = os.path.dirname(x)
        try:
            df = pd.read_excel(x, dtype=str).fillna("")
        except Exception:
            continue
        colmap={}
        for c in df.columns:
            k=str(c).strip()
            k_clean = re.sub(r"[\s\u3000]+", "", k)
            k_clean = re.sub(r"[（(].*?[)）]", "", k_clean)
            if k in ["序号","编号","合同编号"]: colmap[c]="序号" if "序号" in df.columns else "合同编号"
            if k in ["工程地点及内容","工程名称","项目名称","工程地点"]: colmap[c]="工程地点及内容"
            if k in ["单位名称","单位","甲方","客户名称"]: colmap[c]="单位名称"
            if k in ["签订日期","日期","签署日期"]: colmap[c]="签订日期"
            if k in ["合同额","金额","合同金额"]: colmap[c]="合同额"
            if k in ["结算值","结算金额"]: colmap[c]="结算值"
            if k in ["已付款","已支付","已付金额"]: colmap[c]="已付款"
            if k in ["欠付款","欠付","欠付金额"]: colmap[c]="欠付款"
            if "结算" in k_clean and any(token in k_clean for token in ["值","价","金额"]):
                colmap[c]="结算值"
            if "已付" in k_clean and not any(token in k_clean for token in ["时间","日期"]):
                colmap[c]="已付款"
            if "欠付" in k_clean and not any(token in k_clean for token in ["时间","日期"]):
                colmap[c]="欠付款"
        if colmap: df = df.rename(columns=colmap)

        for row_idx, (_, r) in enumerate(df.iterrows(), start=2):
            item={k:str(r.get(k,"")).strip() for k in ["序号","工程地点及内容","单位名称","签订日期","合同额","结算值","已付款","欠付款","合同编号"] if k in df.columns}
            item["签订日期_norm"] = _norm_date(item.get("签订日期",""))
            item["__source_file"] = x
            item["__row_index"] = row_idx

            base = (item.get("序号") or item.get("合同编号","") or "").strip()
            if not base:
                continue
            pdf_guess = _find_pdf(scan_root, base, CONFIG.get("allowed_exts", [".pdf"]), CONFIG.get("pdf_subdirs", ["DOCS","docs"]))

            item["pdf_path"] = ""; item["pdf_dl"] = ""
            if pdf_guess:
                public_base = CONFIG.get("public_base","/data/contracts")
                prev = CONFIG.get("preview_prefix","/files/")
                down = CONFIG.get("download_prefix","/dl/")
                try:
                    rel = os.path.relpath(pdf_guess, public_base).replace("\\","/")
                    if not rel.startswith("../"):
                        try:
                            st = os.stat(pdf_guess)
                            mns = int(getattr(st, 'st_mtime_ns', int(st.st_mtime*1e9)))
                            qs = f"?v={mns}"
                        except Exception:
                            qs = ""
                        item["pdf_path"] = prev + rel + qs
                        item["pdf_dl"]   = down + rel + qs
                except Exception:
                    pass
            dedup_key = base
            existing_idx = rows_by_key.get(dedup_key)
            if existing_idx is not None:
                existing = rows[existing_idx]
                if existing.get("pdf_path") and not item.get("pdf_path"):
                    continue
                if item.get("pdf_path") and not existing.get("pdf_path"):
                    rows[existing_idx] = item
                continue
            rows_by_key[dedup_key] = len(rows)
            rows.append(item)

    with _ROWS_LOCK:
        _ROWS_CACHE["sig"]  = sig
        _ROWS_CACHE["rows"] = list(rows)
    return rows

def _pending_snapshot_path() -> str:
    data_dir = _resolve_data_dir()
    return os.path.join(data_dir, "pending_snapshot.json")

def _load_pending_snapshot() -> Dict[str, Dict[str, int]]:
    path = _pending_snapshot_path()
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            return data
    except FileNotFoundError:
        return {}
    except Exception:
        return {}
    return {}

def _save_pending_snapshot(snapshot: Dict[str, Dict[str, int]]) -> None:
    path = _pending_snapshot_path()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(snapshot, f, ensure_ascii=False, indent=2)

def _scan_pending_files(pending_dir: str) -> Dict[str, Dict[str, int]]:
    snapshot: Dict[str, Dict[str, int]] = {}
    try:
        for name in os.listdir(pending_dir):
            path = os.path.join(pending_dir, name)
            if not os.path.isfile(path):
                continue
            try:
                st = os.stat(path)
            except Exception:
                continue
            snapshot[name] = {
                "mtime_ns": int(getattr(st, "st_mtime_ns", int(st.st_mtime * 1e9))),
                "size": int(st.st_size),
            }
    except FileNotFoundError:
        return {}
    return snapshot

def _copy_pending_changes() -> None:
    pending_dir = PENDING_DIR
    target_dir = ADD_PENDING_DIR
    if not os.path.isdir(pending_dir):
        _auto_update_log(f"pending 目录不存在: {pending_dir}")
        return
    os.makedirs(target_dir, exist_ok=True)

    previous = _load_pending_snapshot()
    current = _scan_pending_files(pending_dir)
    changed = []
    for name, meta in current.items():
        prev = previous.get(name)
        if not prev or prev.get("mtime_ns") != meta.get("mtime_ns") or prev.get("size") != meta.get("size"):
            changed.append(name)

    if not changed:
        _auto_update_log("pending 无新增或变更文件")
        _save_pending_snapshot(current)
        return

    for name in changed:
        src = os.path.join(pending_dir, name)
        dest = os.path.join(target_dir, name)
        try:
            shutil.copy2(src, dest)
            _auto_update_log(f"复制完成: {src} -> {dest}")
        except Exception as exc:
            _auto_update_log(f"复制失败: {src} -> {dest} ({exc})")

    _save_pending_snapshot(current)

def _run_update_script() -> None:
    try:
        proc = subprocess.run(
            ["/root/pdfsearch/bin/update_all.sh"],
            capture_output=True,
            text=True,
            timeout=1200,
        )
        _auto_update_log(f"更新脚本完成: code={proc.returncode}")
        if proc.stdout:
            _auto_update_log(f"stdout: {proc.stdout.strip()}")
        if proc.stderr:
            _auto_update_log(f"stderr: {proc.stderr.strip()}")
    except Exception as exc:
        _auto_update_log(f"更新脚本失败: {exc}")

def _seconds_until_next_update_window(now: datetime.datetime) -> int:
    today_2300 = now.replace(hour=23, minute=0, second=0, microsecond=0)
    today_2330 = now.replace(hour=23, minute=30, second=0, microsecond=0)

    if now < today_2300:
        target = today_2300
    elif now < today_2330:
        target = today_2330
    else:
        target = today_2300 + datetime.timedelta(days=1)

    return max(30, int((target - now).total_seconds()))


def _auto_update_loop() -> None:
    while True:
        sleep_seconds = 300
        try:
            if _get_auto_update_enabled():
                now = datetime.datetime.now()
                today = now.strftime("%Y-%m-%d")
                state = _load_run_state()

                if now.hour == 23 and state.get("copy_date") != today:
                    _auto_update_log("开始检查 pending 目录变更")
                    _copy_pending_changes()
                    state["copy_date"] = today
                    _save_run_state(state)

                if now.hour == 23 and now.minute >= 30 and state.get("update_date") != today:
                    _auto_update_log("开始执行程序更新")
                    _run_update_script()
                    state["update_date"] = today
                    _save_run_state(state)

                sleep_seconds = _seconds_until_next_update_window(now)
        except Exception as exc:
            _auto_update_log(f"自动更新任务异常: {exc}")
        time.sleep(sleep_seconds)
        # 改为每 5 分钟检查一次，降低无效唤醒。
        time.sleep(300)

class QueryIn(BaseModel):
    工程地点及内容: Optional[str]=""
    单位名称: Optional[str]=""
    签订方式: Optional[str]=""  # 预留
    合同编号: Optional[str]=""
    签订日期: Optional[str]=""   # 支持 年/年月/年月日
    合同额: Optional[str]=""      # 数值等价；"0" 视为忽略
    欠付款为0: Optional[bool]=True
    欠付款不为0: Optional[bool]=True
    # 可选覆盖开关（不传则使用全局配置）
    case_insensitive: Optional[bool]=None
    text_logic_or: Optional[bool]=None
    offset: Optional[int]=0
    limit: Optional[int]=50

class ReloadIn(BaseModel):
    roots: Optional[List[str]] = None
    excel_patterns: Optional[List[str]] = None
    allowed_exts: Optional[List[str]] = None
    public_base: Optional[str] = None
    preview_prefix: Optional[str] = None
    download_prefix: Optional[str] = None
    pdf_subdirs: Optional[List[str]] = None
    case_insensitive: Optional[bool] = None
    amount_numeric_equivalence: Optional[bool] = None
    amount_zero_means_empty: Optional[bool] = None
    text_logic_or: Optional[bool] = None


def _parse_amount_decimal(value) -> decimal.Decimal:
    if value is None:
        return decimal.Decimal("0")
    text = str(value).strip()
    if text == "":
        return decimal.Decimal("0")
    t = (text.replace("￥", "").replace("¥", "").replace("人民币", "").replace("元", "")
             .replace("CNY", "").replace("RMB", "").replace(",", "").replace(" ", ""))
    t = re.sub(r"[^0-9.\-]", "", t)
    if t in ("", "-", ".", "-.", ".-"):
        return decimal.Decimal("0")
    try:
        return decimal.Decimal(t)
    except Exception:
        return decimal.Decimal("0")


def _is_settled_row(row: dict) -> bool:
    paid = _parse_amount_decimal(row.get("已付款", ""))
    contract = _parse_amount_decimal(row.get("合同额", ""))
    settle = _parse_amount_decimal(row.get("结算值", ""))
    unpaid = _parse_amount_decimal(row.get("欠付款", ""))
    paid_matches = (paid != 0 and contract != 0 and paid == contract) or (paid != 0 and settle != 0 and paid == settle)
    return paid_matches and unpaid == 0


def _expand_two_digit_year(year_text: str) -> Optional[int]:
    t = (year_text or "").strip()
    if not t.isdigit():
        return None
    if len(t) == 4:
        y = int(t)
    elif len(t) == 2:
        y = 2000 + int(t)
    else:
        return None
    if 1900 <= y <= 2100:
        return y
    return None


def _normalize_year_ranges(years: List[int]) -> str:
    if not years:
        return ""
    sorted_years = sorted(set(years))
    ranges = []
    start = end = sorted_years[0]
    for y in sorted_years[1:]:
        if y == end + 1:
            end = y
            continue
        ranges.append((start, end))
        start = end = y
    ranges.append((start, end))
    parts = [str(a) if a == b else f"{a}-{b}" for a, b in ranges]
    return ",".join(parts)


def _parse_year_filter_expr(expr: str) -> Optional[Dict]:
    raw = (expr or "").strip()
    if not raw:
        return None
    # 保留原有日期能力：单值且形如 YYYY-MM/ YYYY-M 的输入优先按年月处理
    if "," not in raw and "，" not in raw:
        ym = re.match(r"^\s*(\d{4})\s*-\s*(\d{1,2})\s*$", raw)
        if ym and 1 <= int(ym.group(2)) <= 12:
            return None
    tokens = [t.strip() for t in re.split(r"[,，]+", raw) if t.strip()]
    if not tokens:
        return None

    years = set()
    for token in tokens:
        if "-" in token:
            segs = [s.strip() for s in token.split("-")]
            if len(segs) != 2:
                return None
            start = _expand_two_digit_year(segs[0])
            end = _expand_two_digit_year(segs[1])
            if start is None or end is None:
                return None
            if start > end:
                start, end = end, start
            years.update(range(start, end + 1))
        else:
            y = _expand_two_digit_year(token)
            if y is None:
                return None
            years.add(y)

    if not years:
        return None

    normalized = _normalize_year_ranges(sorted(years))
    span_count = len(normalized.split(",")) if normalized else 0
    return {
        "years": years,
        "normalized": normalized,
        "continuous": span_count <= 1,
    }


def _collect_search_results(q: QueryIn):
    # 读取配置/覆盖
    ci = q.case_insensitive if q.case_insensitive is not None else CONFIG.get("case_insensitive", True)
    text_or = q.text_logic_or if q.text_logic_or is not None else CONFIG.get("text_logic_or", False)
    amt_numeric = CONFIG.get("amount_numeric_equivalence", True)

    def norm_text(s: str) -> str:
        s = str(s or "")
        return s.lower() if ci else s

    data = _load_all_rows()
    res = []

    kw_loc = norm_text((q.工程地点及内容 or "").strip())
    kw_unit = norm_text((q.单位名称 or "").strip())
    kw_no_raw = norm_text((q.合同编号 or "").strip())
    kw_no_terms = [term.strip() for term in kw_no_raw.replace("，", ",").split(",") if term.strip()]
    raw_date_input = (q.签订日期 or "").strip()
    year_filter = _parse_year_filter_expr(raw_date_input)
    kw_date = ""
    if not year_filter:
        kw_date = _norm_in_date_std(raw_date_input)
        if isinstance(kw_date, str) and kw_date.isdigit() and len(kw_date) == 4:
            kw_date = kw_date + "-"   # 年/年月/年月日 → 前缀匹配
    kw_amt = _normalize_amount_to_decimal((q.合同额 or "").strip()) if amt_numeric else None

    include_unpaid_zero = q.欠付款为0 if q.欠付款为0 is not None else True
    include_unpaid_non_zero = q.欠付款不为0 if q.欠付款不为0 is not None else True

    # 哪些文本条件参与（空的不参与）
    text_filters = []
    if kw_loc:
        text_filters.append(("工程地点及内容", kw_loc))
    if kw_unit:
        text_filters.append(("单位名称", kw_unit))
    if kw_no_terms:
        text_filters.append(("合同编号_or_序号", kw_no_terms))
    if year_filter:
        text_filters.append(("签订年份", year_filter.get("years", set())))
    elif kw_date:
        text_filters.append(("签订日期_norm_prefix", kw_date))

    for it in data:
        ok = True

        # 1) 金额数值等价（提供且非0时才参与）
        if amt_numeric and (kw_amt is not None):
            item_amt = _normalize_amount_to_decimal(it.get("合同额", ""))
            if item_amt is None or item_amt != kw_amt:
                ok = False
        if not ok:
            continue

        # 2) 文本条件：AND（默认）；OR 可通过 text_logic_or=true 切换
        if text_filters:
            hits = []
            for kind, val in text_filters:
                if kind == "工程地点及内容":
                    hits.append(val in norm_text(it.get("工程地点及内容", "")))
                elif kind == "单位名称":
                    hits.append(val in norm_text(it.get("单位名称", "")))
                elif kind == "合同编号_or_序号":
                    hay = norm_text(it.get("合同编号", "") or it.get("序号", ""))
                    hits.append(any(term in hay for term in val))
                elif kind == "签订日期_norm_prefix":
                    cur = _norm_in_date(it.get("签订日期_norm", "") or it.get("签订日期", ""))
                    hits.append(globals().get('_date_match', _date_match)(kw_date, cur))
                elif kind == "签订年份":
                    cur = _norm_in_date_std(it.get("签订日期_norm", "") or it.get("签订日期", ""))
                    row_year = int(cur[:4]) if len(cur) >= 4 and cur[:4].isdigit() else None
                    hits.append(row_year in val if row_year is not None else False)
            if text_or:
                if not any(hits):
                    ok = False
            else:
                if not all(hits):
                    ok = False

        if not ok:
            continue

        is_settled = _is_settled_row(it)
        if (is_settled and not include_unpaid_zero) or ((not is_settled) and not include_unpaid_non_zero):
            continue

        item = {k: it.get(k, "") for k in RETURN_FIELDS}
        item["__source_file"] = it.get("__source_file", "")
        item["__row_index"] = it.get("__row_index", 0)
        res.append(item)

    return res, kw_date, year_filter

class AutoUpdateToggleIn(BaseModel):
    enabled: bool
    password: Optional[str] = ""

class EntryDetailIn(BaseModel):
    source_file: str
    row_index: int

class EntryRemarkIn(BaseModel):
    source_file: str
    row_index: int
    remark: Optional[str] = ""

@app.get("/api/health")
def health(): return {"ok": True, "ts": int(_utc_now_ts())}

@app.post("/api/mdirs/login")
def login(data: dict):
    if not data or not _password_valid(data.get("password", "") or data.get("pwd", "")):
        raise HTTPException(401, "Bad password")
    return {"token": _issue_token(24), "expires_in_hours": 24}

@app.post("/api/mdirs/reload", dependencies=[Depends(require_auth)])
def reload_cfg(body: ReloadIn, x_auth: str = Header(None)):
    with CONFIG_LOCK:
        for k,v in body.model_dump(exclude_none=True).items():
            if k in CONFIG: CONFIG[k]=v
    return {"ok": True, "config": CONFIG}

@app.post("/api/search", dependencies=[Depends(require_auth)])
def search(q: QueryIn, x_auth: str = Header(None)):
    res, kw_date, year_filter = _collect_search_results(q)

    off=max(0, int(q.offset or 0)); lim=min(200, max(1, int(q.limit or 50)))
    return {"count": len(res), "count_strict": sum(1 for _it in res if str(_it.get("序号","")).strip()), "items": res[off:off+lim], "offset": off, "limit": lim, "debug": {"kw_date": kw_date, "year_filter": (year_filter or {}).get("normalized", ""), "sample_cur": [(_norm_in_date(it.get("签订日期_norm","") or it.get("签订日期",""))) for it in res[:5]]}}  # DEBUG_DATE_SNIPPET


@app.post("/api/search/export", dependencies=[Depends(require_auth)])
def search_export(q: QueryIn):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    rows, kw_date, year_filter = _collect_search_results(q)

    wb = Workbook()
    ws = wb.active
    ws.title = "搜索结果"

    headers = ["序号", "工程地点及内容", "单位名称", "签订日期", "合同额", "结算值", "已付款", "欠付款", "合同编号"]
    ws.append(headers)

    font_header = Font(name="Microsoft YaHei", size=11, bold=True)
    font_body = Font(name="Microsoft YaHei", size=10)
    align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="top", wrap_text=True)
    fill_settled = PatternFill(fill_type="solid", fgColor="EAF6EE")

    ws.row_dimensions[1].height = 24
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = font_header
        cell.alignment = align_header

    col_widths = {
        "A": 14,
        "B": 40,
        "C": 24,
        "D": 14,
        "E": 15,
        "F": 15,
        "G": 15,
        "H": 15,
        "I": 22,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    for idx, item in enumerate(rows, start=2):
        values = [item.get(k, "") for k in headers]
        ws.append(values)
        ws.row_dimensions[idx].height = 36
        settled = _is_settled_row(item)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=idx, column=col_idx)
            cell.font = font_body
            if col_idx in (5, 6, 7, 8):
                cell.alignment = align_right
            elif col_idx in (1, 4):
                cell.alignment = align_center
            else:
                cell.alignment = align_left
            if settled:
                cell.fill = fill_settled

    summary = wb.create_sheet("汇总")
    summary.column_dimensions["A"].width = 18
    summary.column_dimensions["B"].width = 20
    summary.column_dimensions["C"].width = 18
    summary.column_dimensions["D"].width = 18
    summary.column_dimensions["E"].width = 18

    summary_headers = ["分类", "合同总数", "欠付款总数", "起始日期", "结束日期"]
    summary.append(summary_headers)
    for cidx in range(1, len(summary_headers) + 1):
        c = summary.cell(row=1, column=cidx)
        c.font = font_header
        c.alignment = align_header
    summary.row_dimensions[1].height = 24

    def classify_party(index_no: str) -> str:
        t = (index_no or "").upper()
        if "GF" in t:
            return "国丰"
        if "HT" in t:
            return "华腾"
        if "DQ" in t:
            return "蝶泉"
        return "其他"

    parties = {"总计": rows, "国丰": [], "华腾": [], "蝶泉": [], "其他": []}
    for row in rows:
        parties[classify_party(str(row.get("序号", "")))].append(row)

    date_values = [_norm_date(str(r.get("签订日期", ""))) for r in rows]
    date_values = [d for d in date_values if d]
    date_start = min(date_values) if date_values else (kw_date or "-")
    date_end = max(date_values) if date_values else (kw_date or "-")
    if year_filter and not year_filter.get("continuous", True):
        date_start = f"按年份筛选: {year_filter.get('normalized', '')}"
        date_end = "非连续年份"

    summary_rows = []
    for name in ["总计", "国丰", "华腾", "蝶泉", "其他"]:
        group_rows = parties[name]
        unpaid_sum = sum((_parse_amount_decimal(r.get("欠付款", "")) for r in group_rows), decimal.Decimal("0"))
        summary_rows.append([name, len(group_rows), f"{unpaid_sum:.2f}", date_start or "-", date_end or "-"])

    for row in summary_rows:
        summary.append(row)

    for ridx in range(2, 2 + len(summary_rows)):
        summary.row_dimensions[ridx].height = 22
        for cidx in range(1, 6):
            c = summary.cell(row=ridx, column=cidx)
            c.font = font_body
            c.alignment = align_center if cidx != 3 else align_right

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"搜索结果导出_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = quote(filename)
    headers_resp = {"Content-Disposition": f"attachment; filename=export.xlsx; filename*=UTF-8''{encoded_filename}"}
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers_resp)

def _resolve_source_file(source_file: str) -> str:
    if not source_file:
        raise HTTPException(status_code=400, detail="Missing source file")
    real = os.path.realpath(source_file)
    roots = [os.path.realpath(p) for p in CONFIG.get("roots", []) if p]
    if roots and not any(real == r or real.startswith(r + os.sep) for r in roots):
        raise HTTPException(status_code=403, detail="Invalid source file")
    patterns = [p.lower() for p in CONFIG.get("excel_patterns", []) if p]
    if patterns and os.path.basename(real).lower() not in patterns:
        raise HTTPException(status_code=403, detail="Invalid source file")
    if not os.path.isfile(real):
        raise HTTPException(status_code=404, detail="Source file not found")
    return real

def _remark_is_meaningful(value: str) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    if not text:
        return False
    cleaned = re.sub(r"[\\s,，、;；/|]+", "", text)
    return bool(cleaned)

@app.post("/api/entry/detail", dependencies=[Depends(require_auth)])
def entry_detail(body: EntryDetailIn):
    from openpyxl import load_workbook

    real = _resolve_source_file(body.source_file)
    row_index = int(body.row_index or 0)
    if row_index < 2:
        raise HTTPException(status_code=400, detail="Invalid row index")

    wb = load_workbook(real, read_only=True, data_only=True)
    ws = wb.active
    if row_index > ws.max_row:
        raise HTTPException(status_code=404, detail="Row not found")

    headers = []
    for cell in ws[1]:
        val = "" if cell.value is None else str(cell.value).strip()
        headers.append(val)

    row_values = None
    for row in ws.iter_rows(min_row=row_index, max_row=row_index, values_only=True):
        row_values = row
        break
    if row_values is None:
        raise HTTPException(status_code=404, detail="Row not found")

    fields = []
    remark_value = ""
    has_remark_column = False
    for idx, header in enumerate(headers):
        if not header:
            continue
        value = ""
        if idx < len(row_values):
            raw = row_values[idx]
            value = "" if raw is None else str(raw).strip()
        if header == "备注":
            has_remark_column = True
            remark_value = value
            continue
        fields.append({"label": header, "value": value})

    return {"ok": True, "fields": fields, "remark": remark_value, "has_remark_column": has_remark_column}

@app.post("/api/entry/remark", dependencies=[Depends(require_auth)])
def entry_remark(body: EntryRemarkIn):
    from openpyxl import load_workbook

    real = _resolve_source_file(body.source_file)
    row_index = int(body.row_index or 0)
    if row_index < 2:
        raise HTTPException(status_code=400, detail="Invalid row index")

    wb = load_workbook(real)
    ws = wb.active
    if row_index > ws.max_row:
        raise HTTPException(status_code=404, detail="Row not found")

    headers = []
    for cell in ws[1]:
        headers.append("" if cell.value is None else str(cell.value).strip())

    remark_idx = None
    for idx, header in enumerate(headers, start=1):
        if header == "备注":
            remark_idx = idx
            break
    if remark_idx is None:
        remark_idx = len(headers) + 1
        ws.cell(row=1, column=remark_idx, value="备注")

    new_remark = (body.remark or "").strip()
    if not _remark_is_meaningful(new_remark):
        return {"ok": True, "updated": False}

    existing_value = ws.cell(row=row_index, column=remark_idx).value
    existing_remark = "" if existing_value is None else str(existing_value).strip()

    if _remark_is_meaningful(existing_remark):
        if existing_remark == new_remark:
            return {"ok": True, "updated": False}
        if new_remark in existing_remark:
            return {"ok": True, "updated": False}
        combined = f"{existing_remark},{new_remark}"
        ws.cell(row=row_index, column=remark_idx, value=combined)
    else:
        ws.cell(row=row_index, column=remark_idx, value=new_remark)

    wb.save(real)

    with _ROWS_LOCK:
        _ROWS_CACHE["sig"] = None
        _ROWS_CACHE["rows"] = []

    return {"ok": True}

@app.get("/api/auto-update/status", dependencies=[Depends(require_auth)])
def auto_update_status():
    return {"enabled": _get_auto_update_enabled()}

@app.post("/api/auto-update/toggle", dependencies=[Depends(require_auth)])
def auto_update_toggle(body: AutoUpdateToggleIn):
    pw = (body.password or "").strip().lower()
    if pw != "shigeba":
        raise HTTPException(status_code=403, detail="Forbidden")
    _set_auto_update_enabled(body.enabled)
    _auto_update_log(f"自动更新开关切换为: {'开启' if body.enabled else '关闭'}")
    return {"ok": True, "enabled": _get_auto_update_enabled()}

# == injected helpers ==

def _norm_in_date(s: str) -> str:
    """规范用户输入日期：把 年/月/日、点号、斜杠、下划线、空格统一为连字符，再交给 _norm_date"""
    if s is None: return ""
    t = str(s).strip().replace("年","-").replace("月","-").replace("日","")
    import re as _re
    t = _re.sub(r"[._/\\\s]+","-", t)
    return _norm_date(t)


def _date_match(val: str, cur: str) -> bool:
    """val: YYYY / YYYY-MM / YYYY-MM-DD；cur: 任意形态日期字符串"""
    if not val:
        return True
    if not isinstance(val, str):
        val = str(val)
    # 统一记录侧
    try:
        cur = _norm_in_date(cur or "")
    except Exception:
        cur = str(cur or "")
    L = len(val)
    # 显式年前缀 'YYYY-'
    if L == 5 and val.endswith('-'):
        return len(cur) >= 4 and cur[:4] == val[:4]
    # 纯年 'YYYY'
    if L == 4 and val.isdigit():
        return len(cur) >= 4 and cur[:4] == val
    # 年月 'YYYY-MM'
    if L == 7 and val[4] == '-':
        return len(cur) >= 7 and cur[:7] == val
    # 年月日 'YYYY-MM-DD'
    if L == 10 and val[4]=='-' and val[7]=='-':
        return cur == val
    # 兜底：前缀
    return cur.startswith(val)

# ==== canonical date helpers (override) ====
def _date_match(val: str, cur: str) -> bool:
    """
    val: YYYY / YYYY-MM / YYYY-MM-DD / 'YYYY-'(年前缀)
    cur: 任意形态日期（将被规范）
    """
    if not val:
        return True
    if not isinstance(val,str):
        val = str(val)
    try:
        cur = _norm_in_date(cur or "")
    except Exception:
        cur = str(cur or "")
    L = len(val)
    # 显式年前缀 'YYYY-'
    if L==5 and val.endswith('-'):
        return len(cur)>=4 and cur[:4]==val[:4]
    # 纯年 'YYYY'
    if L==4 and val.isdigit():
        return len(cur)>=4 and cur[:4]==val
    # 年月 'YYYY-MM'
    if L==7 and val[4]=='-':
        return len(cur)>=7 and cur[:7]==val
    # 年月日 'YYYY-MM-DD'
    if L==10 and val[4]=='-' and val[7]=='-':
        return cur==val
    # 兜底：前缀
    return cur.startswith(val)

# ==== canonical date std helpers (appended) ====
import re as _re

def _norm_in_date_std(s: str) -> str:
    """统一用户/记录日期为以下三档：'YYYY-'(年前缀) / 'YYYY-MM' / 'YYYY-MM-DD'（补零）"""
    if s is None:
        return ""
    t = str(s).strip().replace("年","-").replace("月","-").replace("日","")
    t = _re.sub(r"[._/\\\s]+", "-", t)  # 统一分隔符
    m = _re.match(r"^(\d{4})(?:-(\d{1,2}))?(?:-(\d{1,2}))?$", t)
    if not m:
        # 兜底：落回旧的 _norm_in_date / _norm_date 流程（若存在）
        try:
            return _norm_in_date(t)  # 兼容旧函数
        except Exception:
            try:
                return _norm_date(t)    # 最后兜底
            except Exception:
                return t
    y,mm,dd = m.groups()
    if not mm:
        return y + "-"                 # 只有年 → 'YYYY-'
    mm = ("0"+mm)[-2:]
    if not dd:
        return f"{y}-{mm}"            # 年月
    dd = ("0"+dd)[-2:]
    return f"{y}-{mm}-{dd}"           # 年月日

def _date_match(val: str, cur: str) -> bool:
    """val: YYYY / YYYY- / YYYY-MM / YYYY-MM-DD；cur: 任意形态，先标准化再比对"""
    if not val:
        return True
    if not isinstance(val,str):
        val = str(val)
    cur = _norm_in_date_std(cur or "")
    L = len(val)
    if L==5 and val.endswith('-'):           # 'YYYY-'
        return len(cur)>=4 and cur[:4]==val[:4]
    if L==4 and val.isdigit():               # 'YYYY'
        return len(cur)>=4 and cur[:4]==val
    if L==7 and val[4]=='-':                 # 'YYYY-MM'
        return len(cur)>=7 and cur[:7]==_norm_in_date_std(val)
    if L==10 and val[4]=='-' and val[7]=='-':# 'YYYY-MM-DD'
        return cur==_norm_in_date_std(val)
    return cur.startswith(val)               # 兜底

# ==== maint endpoint: manual append to*.pdf (one-time secret) ====
from typing import Optional
from pydantic import BaseModel
from fastapi import HTTPException
import subprocess

class MaintAuth(BaseModel):
    secret: Optional[str] = ""

@app.post("/api/maint/append")
def maint_append(a: MaintAuth):
    """
    手动维护：运行 to*.pdf 追加脚本。
    仅校验一次性口令（大小写不敏感）：shigeba
    """
    pw = (a.secret or "").strip().lower()
    if pw != "shigeba":
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        proc = subprocess.run(
            ["/root/pdfsearch/bin/update_all.sh"],
            capture_output=True, text=True, timeout=600
        )
        return {
            "ok": proc.returncode == 0,
            "code": proc.returncode,
            "stdout": (proc.stdout or "").strip(),
            "stderr": (proc.stderr or "").strip()
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))












@app.get("/api/entries/count")
def entries_count():
    from pathlib import Path
    from openpyxl import load_workbook

    STD_HEADER = ["序号","工程地点及内容","单位名称","签订途径","启动时间","结果确定时间","签订日期","控制价","合同额","结算值","已付款","欠付款","备注"]

    def trim(row):
        vals=[("" if v is None else str(v).strip()) for v in row]
        while vals and vals[-1]=="":
            vals.pop()
        return vals

    root = Path("/data/contracts")
    per, total_all, total_strict = [], 0, 0

    years = sorted([p for p in root.iterdir() if p.is_dir() and p.name.isdigit()], key=lambda x:int(x.name))
    for y in years:
        idx = None
        for f in y.iterdir():
            if f.is_file() and f.name.lower()=="index.xlsx":
                idx=f; break
        if not idx:
            continue
        try:
            wb=load_workbook(idx, read_only=True, data_only=True); ws=wb.active
            hdr = trim([c.value for c in ws[1]])
            strict = (len(hdr)==len(STD_HEADER) and all(hdr[i]==STD_HEADER[i] for i in range(len(STD_HEADER))))
            n = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0] not in (None, ""))
            total_all += n
            if strict:
                total_strict += n
            per.append({"year": int(y.name), "rows": n, "strict": strict, "file": str(idx)})
        except Exception as e:
            per.append({"year": int(y.name), "error": str(e), "file": str(idx)})

    return {"ok": True, "total_all": total_all, "total_strict": total_strict, "per_year": per}



@app.post("/do_login")
def do_login(password: str = Form(...)):
    if not _password_valid(password):
        raise HTTPException(401, "Bad password")
    t = _issue_token(24)
    resp = RedirectResponse(url="/", status_code=302)
    resp.set_cookie("X-Auth", t, httponly=True, samesite="lax", max_age=24*3600)
    return resp


AUTO_UPDATE_ENABLED = _load_auto_update_enabled()
threading.Thread(target=_auto_update_loop, daemon=True).start()
