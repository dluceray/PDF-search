import re, pathlib, shutil, sys, os
p=pathlib.Path("/root/pdfsearch/ingest_excels.py")
s=p.read_text(encoding="utf-8")

new_func = r'''
def backup_and_append(idx_path: Path, rows):
    """
    写入策略（占空+尾追加）：
    - 优先把记录写入 A列(序号)为空的既有行（即便该行有边框/样式，也视为空槽）；
    - 空槽用尽后，再使用 ws.append() 追加到表尾。
    """
    from openpyxl import load_workbook
    bak = idx_path.with_name(idx_path.name + f".bak.{ts()}")
    tmp = idx_path.with_suffix(".tmp.xlsx")
    shutil.copy2(idx_path, bak)

    wb = load_workbook(idx_path)
    ws = wb.active

    def a_col_empty(r):
        return (ws.cell(row=r, column=1).value in (None, ""))

    r_ptr = 2
    maxr = ws.max_row  # 含已画表格的空行
    for rec in rows:
        while r_ptr <= maxr and not a_col_empty(r_ptr):
            r_ptr += 1
        if r_ptr <= maxr:
            for ci, val in enumerate(rec, start=1):
                ws.cell(row=r_ptr, column=ci, value=val)
            r_ptr += 1
        else:
            ws.append(rec)

    wb.save(tmp)
    os.replace(tmp, idx_path)
    return bak.name
'''

pat = re.compile(r"def\\s+backup_and_append\\s*\\(.*?\\):[\\s\\S]*?\\n(?=def\\s+|$)")
m = pat.search(s)
if not m:
    print("NOT_FOUND: backup_and_append; appending new def at EOF", file=sys.stderr)
    s = s.rstrip()+"\\n\\n"+new_func.lstrip()
else:
    s = s[:m.start()] + new_func + s[m.end():]

bk = p.with_suffix(".bak.patch."+str(os.getpid()))
shutil.copy2(p, bk)
p.write_text(s, encoding="utf-8")
print("OK: patched", p, "backup:", bk)
