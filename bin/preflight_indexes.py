#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import json, sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path("/data/contracts")
STD_HEADER = ["序号","工程地点及内容","单位名称","签订途径","启动时间","结果确定时间","签订日期","控制价","合同额","结算值","已付款","欠付款","备注"]
HEADER_ALIASES = {
    "序号": ["序号", "编号"],
    "工程地点及内容": ["工程地点及内容", "工程地点", "工程内容"],
    "单位名称": ["单位名称", "单位"],
    "签订途径": ["签订途径", "签订方式"],
    "启动时间": ["启动时间", "启动日期"],
    "结果确定时间": ["结果确定时间", "结果确定日期"],
    "签订日期": ["签订日期", "合同签订日期"],
    "控制价": ["控制价", "控制金额"],
    "合同额": ["合同额", "合同金额"],
    "结算值": ["结算值", "结算金额"],
    "已付款": ["已付款", "已支付", "已付金额"],
    "欠付款": ["欠付款", "欠付", "欠付金额"],
    "备注": ["备注", "备注说明"],
}

def header_equal(a,b): return len(a)==len(b) and all((a[i]==b[i]) for i in range(len(a)))
def match_header(row):
    cells = trim(row)
    for std in STD_HEADER:
        aliases = HEADER_ALIASES.get(std, [std])
        if not any(c in aliases for c in cells):
            return False
    return True

def trim(row):
    vals=[ ("" if v is None else str(v).strip()) for v in row ]
    while vals and vals[-1]=="": vals.pop()
    return vals

report=[]
for y in sorted([p for p in ROOT.iterdir() if p.is_dir() and p.name.isdigit()]):
    idx=None
    for f in y.iterdir():
        if f.is_file() and f.name.lower()=="index.xlsx":
            idx=f; break
    if not idx:
        report.append({"year": y.name, "ok": False, "reason": "缺少index.xlsx"}); continue
    try:
        wb=load_workbook(idx, data_only=True, read_only=True); ws=wb.active
        hdr=trim([c.value for c in ws[1]])
        ok=match_header(hdr)
        report.append({"year": y.name, "ok": ok, "file": str(idx), "header": hdr if not ok else "OK"})
    except Exception as e:
        report.append({"year": y.name, "ok": False, "file": str(idx), "reason": f"读取失败:{e}"})

print(json.dumps({"ok": True, "report": report}, ensure_ascii=False, indent=2))
