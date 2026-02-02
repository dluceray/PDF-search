#!/root/pdfsearch/venv/bin/python

# === probe: log interpreter & xlrd presence ===
try:
    import sys, json, time
    _probe = {"ts": time.strftime("%F %T"), "exe": sys.executable}
    try:
        import xlrd
        _probe["xlrd"] = getattr(xlrd, "__version__", "?")
    except Exception as e:
        _probe["xlrd_error"] = repr(e)
    with open("/tmp/ingest_probe.json","a",encoding="utf-8") as f:
        f.write(json.dumps(_probe, ensure_ascii=False)+"\n")
except Exception:
    pass
# === probe end ===
# -*- coding: utf-8 -*-
import os, sys, glob, time, json, shutil, argparse, re
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Tuple, Optional

YEAR_MIN, YEAR_MAX = 2000, 2050
STD_HEADER = ["序号","工程地点及内容","单位名称","签订途径","启动时间","结果确定时间","签订日期","控制价","合同额","结算值","已付款","欠付款","备注"]
REQUIRED_HEADER = ["序号", "工程地点及内容", "单位名称"]
HEADER_ALIASES = {
    "序号": ["序号", "编号"],
    "工程地点及内容": ["工程地点及内容", "工程地点", "工程内容"],
    "单位名称": ["单位名称", "单位"],
    "签订途径": ["签订途径", "签订方式"],
    "启动时间": ["启动时间", "启动日期"],
    "结果确定时间": ["结果确定时间", "结果确定日期"],
    "签订日期": ["签订日期", "合同签订日期"],
    "控制价": ["控制价", "控制金额"],
    "合同额": ["合同额", "合同金额"],
    "结算值": ["结算值", "结算金额"],
    "已付款": ["已付款", "已支付", "已付金额"],
    "欠付款": ["欠付款", "欠付", "欠付金额"],
    "备注": ["备注", "备注说明"],
}

def ts(): return datetime.now().strftime("%Y%m%d%H%M%S")
def header_equal(a,b): return len(a)==len(b) and all((a[i]==b[i]) for i in range(len(a)))
def _ingest_log_path() -> Path:
    return Path("/tmp/ingest_excels.log")
def _log(msg: str) -> None:
    try:
        _ingest_log_path().parent.mkdir(parents=True, exist_ok=True)
        with _ingest_log_path().open("a", encoding="utf-8") as f:
            f.write(f"[{datetime.now().strftime('%F %T')}] {msg}\n")
    except Exception:
        pass
def normalize_header_cells(row):
    vals=[ ("" if v is None else str(v).strip()) for v in row ]
    while vals and vals[-1]=="": vals.pop()
    return vals
def row_has_data(row) -> bool:
    return any((v is not None and str(v).strip() != "") for v in row)
def canonicalize_header(cell: str):
    val = (cell or "").strip()
    if not val:
        return ""
    for std, aliases in HEADER_ALIASES.items():
        if val in aliases:
            return std
    return val
def is_meaningful_header(name: str) -> bool:
    val = (name or "").strip()
    if not val:
        return False
    lower = val.lower()
    if lower in {"nan", "none", "null"}:
        return False
    if lower.startswith("unnamed"):
        return False
    if re.fullmatch(r"\d+", val):
        return False
    if re.fullmatch(r"(列|欄|字段|field|column|col)\s*\d+", val, flags=re.IGNORECASE):
        return False
    if re.fullmatch(r"[^\w\u4e00-\u9fff]+", val):
        return False
    return True

def normalize_merged_cells(ws):
    try:
        merged_ranges = list(ws.merged_cells.ranges)
    except Exception as exc:
        return False, f"读取合并单元格失败: {exc}"
    if not merged_ranges:
        return True, ""
    _log(f"normalize_merged_cells sheet={ws.title} ranges={len(merged_ranges)}")
    for cell_range in merged_ranges:
        try:
            min_row, min_col, max_row, max_col = (
                cell_range.min_row,
                cell_range.min_col,
                cell_range.max_row,
                cell_range.max_col,
            )
            top_left = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(str(cell_range))
            for row_idx in range(min_row, max_row + 1):
                for col_idx in range(min_col, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is None:
                        cell.value = top_left
        except Exception as exc:
            return False, f"处理合并单元格失败({cell_range}): {exc}"
    return True, ""
def header_map_from_cells(cells: List[str]):
    header_map = {}
    headers = []
    for i, cell in enumerate(cells):
        name = canonicalize_header(cell)
        headers.append(name)
        if name and name not in header_map:
            header_map[name] = i
    return header_map, headers
def headers_present_in_rows(rows: List[Dict], headers_seen: List[str]) -> List[str]:
    present = set()
    for row in rows:
        for key in row.keys():
            if not key or not is_meaningful_header(key):
                continue
            present.add(key)
    ordered = []
    for name in headers_seen:
        if name in present and name not in ordered:
            ordered.append(name)
    for name in present:
        if name not in ordered:
            ordered.append(name)
    return ordered
def match_header(row):
    cells = normalize_header_cells(row)
    header_map, headers = header_map_from_cells(cells)
    missing_required = [k for k in REQUIRED_HEADER if k not in header_map]
    if missing_required:
        return False, {}, headers
    return True, header_map, headers
def parse_year_from_seq(s: str):
    s=(s or "").strip()
    if len(s)<2 or not s[:2].isdigit(): return False,-1,"序号前两位非数字"
    yy=int(s[:2]); year=2000+yy
    if not (YEAR_MIN<=year<=YEAR_MAX): return False,-1,f"年份超出范围:{year}"
    return True,year,""

def find_index_file(year_dir: Path) -> Optional[Path]:
    """在年份目录中大小写不敏感地寻找 index.xlsx；返回实际存在的文件路径（保留原始大小写）。"""
    if not year_dir.is_dir(): return None
    for p in year_dir.iterdir():
        if p.is_file() and p.name.lower()=="index.xlsx":
            return p
    return None

def _create_empty_index(idx_path: Path, headers: List[str]) -> Optional[Path]:
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(idx_path)
        return idx_path
    except Exception:
        return None

def build_required_headers() -> List[str]:
    headers = []
    for name in ["序号", *REQUIRED_HEADER]:
        cname = canonicalize_header(name)
        if cname and cname not in headers:
            headers.append(cname)
    return headers

def build_year_headers(headers_needed: List[str]) -> List[str]:
    headers = build_required_headers()
    for name in headers_needed:
        cname = canonicalize_header(name)
        if cname and is_meaningful_header(cname) and cname not in headers:
            headers.append(cname)
    return headers

def ensure_year_index(root: Path, year: int, headers_needed: List[str]) -> Optional[Path]:
    year_dir = root / str(year)
    idx_path = find_index_file(year_dir)
    if idx_path:
        return idx_path
    year_dir.mkdir(parents=True, exist_ok=True)
    idx_path = year_dir / "index.xlsx"
    headers = build_year_headers(headers_needed)
    return _create_empty_index(idx_path, headers)

def load_year_sheet(idx_path: Path, headers_needed: List[str]):
    if not idx_path or not idx_path.is_file():
        return False, None, None, {}, {}, "缺少index.xlsx"
    try:
        from openpyxl import load_workbook
        wb=load_workbook(idx_path); ws=wb.active
        ok, why = normalize_merged_cells(ws)
        if not ok:
            return False, None, None, {}, {}, f"无法读取index.xlsx:{why}"
        hdr=[ ("" if c.value is None else str(c.value).strip()) for c in ws[1] ]
        trimmed_headers = normalize_header_cells(hdr)
        canonical_headers = [canonicalize_header(h) for h in trimmed_headers]
        header_map, _headers = header_map_from_cells(canonical_headers)
        desired_headers = []
        seq_name = canonicalize_header("序号")
        if seq_name:
            desired_headers.append(seq_name)
        for key in headers_needed:
            name = canonicalize_header(key)
            if name and is_meaningful_header(name) and name not in desired_headers:
                desired_headers.append(name)
        for key in canonical_headers:
            if key and is_meaningful_header(key) and key not in desired_headers:
                desired_headers.append(key)
        if desired_headers != canonical_headers:
            old_index = {name: i for i, name in enumerate(canonical_headers) if name}
            max_row = ws.max_row
            for row_idx in range(1, max_row + 1):
                if row_idx == 1:
                    new_values = list(desired_headers)
                else:
                    old_values = [cell.value for cell in ws[row_idx]]
                    new_values = []
                    for name in desired_headers:
                        idx = old_index.get(name)
                        new_values.append(old_values[idx] if idx is not None and idx < len(old_values) else None)
                for col_idx, value in enumerate(new_values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
            if ws.max_column > len(desired_headers):
                ws.delete_cols(len(desired_headers) + 1, ws.max_column - len(desired_headers))
        header_map, _headers = header_map_from_cells(desired_headers)
        seq_pos = header_map.get("序号")
        if seq_pos is None:
            return False, None, None, {}, {}, "索引表缺少必填列: 序号"
        seq_map = {}
        for r in ws.iter_rows(min_row=2):
            if not r:
                continue
            val = r[seq_pos].value if seq_pos < len(r) else None
            if val is None:
                continue
            seq = str(val).strip()
            if not seq:
                continue
            seq_map[seq] = r[0].row
        return True, wb, ws, header_map, seq_map, ""
    except Exception as e:
        return False, None, None, {}, {}, f"无法读取index.xlsx:{e}"


def backup_and_merge(idx_path: Path, rows, headers_needed: List[str], allowed_headers: List[str]):
    """
    写入策略（覆盖+追加）：
    - 若序号已存在，则按导入值覆盖该行对应列；
    - 若序号不存在，则追加新行；
    - 若导入包含服务端缺失且表头有意义的列，则自动新增列。
    """
    date_dir = idx_path.parent / datetime.now().strftime("%Y%m%d")
    date_dir.mkdir(parents=True, exist_ok=True)
    bak = date_dir / idx_path.name
    if bak.exists():
        bak = date_dir / f"{idx_path.stem}.{ts()}{idx_path.suffix}"
    tmp = idx_path.with_suffix(".tmp.xlsx")
    _log(f"backup_and_merge start index={idx_path} rows={len(rows)} headers_needed={headers_needed}")
    shutil.copy2(idx_path, bak)

    ok, wb, ws, header_map, seq_map, why = load_year_sheet(idx_path, headers_needed)
    if not ok:
        raise ValueError(why)
    merged_cell_map = {}
    try:
        for cell_range in ws.merged_cells.ranges:
            min_row, min_col, max_row, max_col = cell_range.min_row, cell_range.min_col, cell_range.max_row, cell_range.max_col
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    if row == min_row and col == min_col:
                        continue
                    merged_cell_map[(row, col)] = (min_row, min_col)
    except Exception:
        merged_cell_map = {}

    allowed = {canonicalize_header(h) for h in allowed_headers if h}
    added = 0
    updated = 0
    for rec in rows:
        seq = str(rec.get("序号", "") or "").strip()
        if not seq:
            continue
        row_idx = seq_map.get(seq)
        if row_idx is None:
            row_idx = ws.max_row + 1
            seq_map[seq] = row_idx
            added += 1
        else:
            updated += 1
        for key, val in rec.items():
            if not key:
                continue
            if key not in header_map:
                if key not in allowed or not is_meaningful_header(key):
                    continue
                col = ws.max_column + 1
                ws.cell(row=1, column=col, value=key)
                header_map[key] = col - 1
            col_idx = header_map[key] + 1
            write_row, write_col = row_idx, col_idx
            target = merged_cell_map.get((write_row, write_col))
            if target:
                write_row, write_col = target
            ws.cell(row=write_row, column=write_col, value=val)

    wb.save(tmp)
    os.replace(tmp, idx_path)
    _log(f"backup_and_merge done index={idx_path} backup={bak} added={added} updated={updated}")
    return bak.name, added, updated
def detect_header_row(values_rows, mode:str):
    rows=[normalize_header_cells(r) for r in values_rows[:5]]
    if mode=="2":
        hdr = rows[1] if len(rows)>1 else []
        ok, header_map, cells = match_header(hdr) if len(rows)>1 else (False, {}, [])
        return (len(rows)>1 and ok, 1, "表头不匹配" if not (len(rows)>1 and ok) else "", cells, header_map)
    if mode=="1":
        hdr = rows[0] if rows else []
        ok, header_map, cells = match_header(hdr) if len(rows)>0 else (False, {}, [])
        return (len(rows)>0 and ok, 0, "表头不匹配" if not (len(rows)>0 and ok) else "", cells, header_map)
    for idx, row in enumerate(rows):
        ok, header_map, cells = match_header(row)
        if ok:
            return True, idx, "", cells, header_map
    return False,-1,"前五行均非标准表头", rows[0] if rows else [], {}

def is_summary_sheet_name(name: str) -> bool:
    if not name:
        return False
    normalized = str(name).strip().lower().replace(" ", "")
    keywords = [
        "汇总",
        "汇總",
        "总结",
        "總結",
        "总表",
        "總表",
        "总览",
        "總覽",
        "总计",
        "總計",
        "合计",
        "合計",
        "统计",
        "統計",
        "summary",
    ]
    return any(k in normalized for k in keywords)

def load_pending_rows(xlsx_path: Path, header_mode: str):
    ext = xlsx_path.suffix.lower()
    errors = []
    rows_out = []
    headers_seen = []
    try:
        _log(f"load_pending_rows start file={xlsx_path} ext={ext} header_mode={header_mode}")
        if ext == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path, data_only=True, read_only=True)
            _log(f"openpyxl loaded sheets={len(wb.worksheets)} names={[ws.title for ws in wb.worksheets]}")
            for ws in wb.worksheets:
                if is_summary_sheet_name(ws.title):
                    _log(f"skip summary sheet={ws.title}")
                    continue
                values_rows = [ (list(r) if r else []) for r in ws.iter_rows(min_row=1, values_only=True) ]
                if not any(any((v is not None and str(v).strip() != "") for v in row) for row in values_rows):
                    _log(f"skip empty sheet={ws.title}")
                    continue
                ok, hdr_idx, why, hdr_cells, header_map = detect_header_row(values_rows, header_mode)
                if not ok:
                    _log(f"sheet={ws.title} header_detect_failed reason={why} first_rows={values_rows[:3]}")
                    errors.append(f"{ws.title}: {why}")
                    continue
                _log(f"sheet={ws.title} header_row_index={hdr_idx} header={hdr_cells}")
                for h in hdr_cells:
                    if h and h not in headers_seen:
                        headers_seen.append(h)
                added_rows = 0
                for r in values_rows[hdr_idx+1:]:
                    r = list(r) if isinstance(r, (list, tuple)) else [r]
                    if not row_has_data(r):
                        continue
                    row = {}
                    for key, ci in header_map.items():
                        row[key] = r[ci] if ci < len(r) else ""
                    rows_out.append(row)
                    added_rows += 1
                _log(f"sheet={ws.title} rows_added={added_rows}")
        elif ext == ".xls":
            try:
                import xlrd  # 1.2.0
            except Exception:
                _log("xls missing xlrd==1.2.0")
                return False, "处理 .xls 需要安装 xlrd==1.2.0", [], []
            book = xlrd.open_workbook(xlsx_path, formatting_info=False)
            _log(f"xlrd loaded sheets={book.nsheets} names={[book.sheet_by_index(i).name for i in range(book.nsheets)]}")
            for si in range(book.nsheets):
                sh = book.sheet_by_index(si)
                if is_summary_sheet_name(sh.name):
                    _log(f"skip summary sheet={sh.name}")
                    continue
                values_rows = [[sh.cell_value(i,j) for j in range(sh.ncols)] for i in range(sh.nrows)]
                if not any(any((v is not None and str(v).strip() != "") for v in row) for row in values_rows):
                    _log(f"skip empty sheet={sh.name}")
                    continue
                ok, hdr_idx, why, hdr_cells, header_map = detect_header_row(values_rows, header_mode)
                if not ok:
                    _log(f"sheet={sh.name} header_detect_failed reason={why} first_rows={values_rows[:3]}")
                    errors.append(f"{sh.name}: {why}")
                    continue
                _log(f"sheet={sh.name} header_row_index={hdr_idx} header={hdr_cells}")
                for h in hdr_cells:
                    if h and h not in headers_seen:
                        headers_seen.append(h)
                added_rows = 0
                for r in values_rows[hdr_idx+1:]:
                    r = list(r) if isinstance(r, (list, tuple)) else [r]
                    if not row_has_data(r):
                        continue
                    row = {}
                    for key, ci in header_map.items():
                        row[key] = r[ci] if ci < len(r) else ""
                    rows_out.append(row)
                    added_rows += 1
                _log(f"sheet={sh.name} rows_added={added_rows}")
        else:
            return False, "不支持的扩展名（仅 .xlsx/.xls）", [], []
    except Exception as e:
        _log(f"load_pending_rows exception={e!r}")
        return False, f"无法读取Excel: {e}", [], []

    if errors:
        _log(f"load_pending_rows failed errors={errors}")
        return False, "以下工作表不符合要求（全部中止处理）：" + "; ".join(errors), [], []
    _log(f"load_pending_rows done file={xlsx_path} rows_out={len(rows_out)} headers_seen={headers_seen}")
    return True, "", rows_out, headers_seen


def process_file(xlsx: Path, root: Path, header_mode: str):
    res={"file": xlsx.name, "added":0, "updated":0, "skipped":0, "invalid":0, "details":[]}
    ok, why, rows, headers_seen = load_pending_rows(xlsx, header_mode)
    _log(f"process_file file={xlsx} ok={ok} rows={len(rows)} headers_seen={headers_seen}")
    if not ok: return {"file": xlsx.name, "__file_failed__": True, "reason": why}
    buckets: Dict[int, List[Dict]] = {}
    for r in rows:
        seq = "" if r.get("序号") is None else str(r.get("序号")).strip()
        if seq=="":
            res["invalid"]+=1; res["details"].append({"seq":None,"status":"invalid","reason":"序号为空"}); continue
        ok_year,year,rr = parse_year_from_seq(seq)
        if not ok_year:
            res["invalid"]+=1; res["details"].append({"seq":seq,"status":"invalid","reason":rr}); continue
        buckets.setdefault(year, []).append(r)
    for year, rows2 in buckets.items():
        if not rows2: continue
        year_headers = build_year_headers(headers_present_in_rows(rows2, headers_seen))
        idx_path = ensure_year_index(root, year, year_headers)
        if not idx_path:
            res["invalid"] += len(rows2)
            res["details"].append({"year": year, "status":"invalid", "reason":"无法创建index.xlsx"})
            continue
        _log(f"process_file file={xlsx} year={year} rows={len(rows2)} index={idx_path}")
        lock = idx_path.with_name(idx_path.name + ".lock")
        waited=0.0
        while lock.exists():
            import time as _t; _t.sleep(0.1); waited+=0.1
            if waited>5.0: return {"file": xlsx.name, "__file_failed__": True, "reason": f"年索引被锁:{idx_path.name}"}
        try:
            _log(f"process_file file={xlsx} year={year} acquiring_lock waited={waited:.1f}s")
            lock.touch()
            start_ts = time.time()
            _log(f"process_file file={xlsx} year={year} merge_start rows={len(rows2)}")
            bak, added, updated = backup_and_merge(idx_path, rows2, year_headers, year_headers)
            _log(f"process_file file={xlsx} year={year} merge_done elapsed={time.time()-start_ts:.2f}s")
            res["details"].append({"year":year,"status":"wrote","rows":len(rows2),"backup":bak})
            res["added"] += added
            res["updated"] += updated
            _log(f"process_file file={xlsx} year={year} added={added} updated={updated} backup={bak}")
        except Exception as exc:
            _log(f"process_file file={xlsx} year={year} merge_failed error={exc!r}")
            raise
        finally:
            try: lock.unlink(missing_ok=True)
            except: pass
    return res

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--contracts-root", default="/data/contracts")
    ap.add_argument("--pending-dir", default="/data/add_pending")
    ap.add_argument("--done-dir", default="/data/add_done")
    ap.add_argument("--error-dir", default="/data/add_error")
    ap.add_argument("--summary-out", default="")
    ap.add_argument("--lock-file", default="/var/run/excel-ingest.lock")
    ap.add_argument("--header-row", default="2", choices=["1","2","auto"], help="待处理 Excel 的表头所在行（1-based），默认 2")
    args=ap.parse_args()
    root=Path(args.contracts_root); pending=Path(args.pending_dir); done=Path(args.done_dir); error=Path(args.error_dir)
    _log(f"ingest start contracts_root={root} pending={pending} done={done} error={error} header_row={args.header_row}")
    if os.path.exists(args.lock_file):
        try:
            raw = Path(args.lock_file).read_text(encoding="utf-8").strip()
            pid = int(raw) if raw else None
        except Exception:
            pid = None
        if pid:
            try:
                os.kill(pid, 0)
                print(json.dumps({"ok": False, "error": "another ingest running"}, ensure_ascii=False))
                sys.exit(1)
            except OSError:
                pass
        try:
            os.unlink(args.lock_file)
        except Exception:
            pass
    try:
        fd=os.open(args.lock_file, os.O_CREAT | os.O_EXCL | os.O_WRONLY, 0o644)
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(str(os.getpid()))
    except FileExistsError:
        print(json.dumps({"ok": False, "error": "another ingest running"}, ensure_ascii=False)); sys.exit(1)
    try:
        files=[p for p in pending.iterdir() if p.is_file() and not p.name.startswith("~$") and p.suffix.lower() in {".xls",".xlsx",".xlsm"}]
        files.sort(key=lambda p: p.stat().st_mtime)
        from pathlib import Path as _P
        _P("/tmp/ingest_scan.json").write_text("\n".join(f.name for f in files)+"\n", encoding="utf-8")
        total={"files_total":len(files),"files_processed":0,"files_failed":0,"rows_added":0,"rows_updated":0,"rows_skipped":0,"rows_invalid":0,"per_file":[]}
        for f in files:
            res=process_file(f, root, args.header_row)
            if res.get("__file_failed__"):
                error.mkdir(parents=True, exist_ok=True); shutil.move(str(f), str(error/f.name))
                if res.get("reason"): (error/(f.name + ".reason.txt")).write_text(res["reason"], encoding="utf-8")
                total["files_failed"]+=1; total["per_file"].append(res); continue
            done.mkdir(parents=True, exist_ok=True); shutil.move(str(f), str(done/f.name))
            total["files_processed"]+=1
            total["rows_added"]+=res["added"]; total["rows_updated"]+=res["updated"]; total["rows_skipped"]+=res["skipped"]; total["rows_invalid"]+=res["invalid"]
            total["per_file"].append(res)
        if args.summary_out:
            Path(args.summary_out).write_text(json.dumps({"ok":True, **total}, ensure_ascii=False, indent=2), encoding="utf-8")
        print(json.dumps({"ok":True, **total}, ensure_ascii=False, indent=2))
    finally:
        try: os.unlink(args.lock_file)
        except: pass

if __name__=="__main__": main()
